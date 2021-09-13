import time
from random import randint
import json
import openpyxl
from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.firefox.options import Options
# from requests.exceptions import HTTPError


# fp = FirefoxProfile(r'C:\Users\gtcan\AppData\Roaming\Mozilla\Firefox\Profiles\26uglqi9.default-beta')
options = Options()
# options.profile = fp
driver = webdriver.Firefox(options=options)
wb = openpyxl.load_workbook(filename='900Startups.xlsx')
ws = wb[wb.sheetnames[0]]
filteredSheet = wb[wb.sheetnames[1]]


def find_data(website):
    driver.get('view-source:https://data.similarweb.com/api/v1/data?domain=' + website)
    pagetxt = driver.find_element_by_tag_name('pre').text
    try:
        if driver.find_elements_by_tag_name('p') != []:
            time.sleep(45)
            pagedata = json.loads(pagetxt)
            traffic = pagedata['EstimatedMonthlyVisits']['2021-02-01']
            time.sleep(randint(4, 10))
            return traffic
        else:
            print(pagetxt)
            pagedata = json.loads(pagetxt)
            traffic = pagedata['EstimatedMonthlyVisits']['2021-02-01']
            time.sleep(randint(4, 10))
            return traffic
    except AttributeError:
        return 0
    except KeyError:
        return 0


def iterator(stopped):
    traffic = find_data(ws["I" + str(stopped)].value)
    if traffic >= 350000:
        for column in "ABCDEFGHIJKLMNO":
            filteredSheet[column + str(stopped)].value = ws[column + str(stopped)].value
        wb.save("900Startups.xlsx")
    print(stopped)


for x in range(767, 814):
    iterator(x)
wb.save("900Startups.xlsx")
